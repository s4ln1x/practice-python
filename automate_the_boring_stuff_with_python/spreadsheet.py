#!/usr/bin/env python3

import openpyxl

wbe = openpyxl.load_workbook('example.xlsx')

# Return a list of all the available sheets
print(f'All the sheets on the example.xlsx file {wbe.sheetnames}')

sheet_1 = wbe['2020']
print('This is the whole column A')
for i in range(1, 57):
    print(sheet_1[f'A{i}'].value)

# Creating a new spreadsheet
# Default name of the new sheet is 'Sheet'
wb = openpyxl.Workbook()

sheet_3 = wb['Sheet']

sheet_3.title = 'This is a better name'

# index is optional
sheet_2 = wb.create_sheet(index=0, title='copy of 2020 pops')


# Copy all the sheet from example sheet 1 to the new workbook sheet_3
columns = 'A B C D E F G H I J K L'
for column in columns.split(' '):
    for row in range(1, 57):
        sheet_3[f'{column}{row}'] = sheet_1[f'{column}{row}'].value

# Add how to manage the sheets with rows and colums too, not just with cells aka A1
for i in range(1, 5):
    for j in range(1, 5):
        print(sheet_2.cell(row=j, column=i).value)

wb.save('example2.xlsx')
